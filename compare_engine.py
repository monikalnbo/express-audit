# -*- coding: utf-8 -*-
"""
快递单号比对审查引擎
功能:
  1. 导入两份 Excel,按"快递单号"匹配
  2. 单边存在的快递单号 -> 异常(缺失)
  3. 重量模糊比对: 一边具体重量 vs 一边区间重量(如 "1-2kg"、"0.5~1.5公斤")
  4. 所有异常标注 + 汇总提取到异常表
"""
import re
import unicodedata
import pandas as pd
from openpyxl.styles import PatternFill, Font

# ---------- 区间重量解析 ----------
UNIT_FACTOR = {"mg": 0.000001, "g": 0.001, "克": 0.001, "kg": 1.0, "公斤": 1.0, "千克": 1.0,
               "吨": 1000.0, "t": 1000.0, "lb": 0.453592, "磅": 0.453592}

def parse_weight_value(text):
    """解析具体重量, 返回 kg 数值; 失败返回 None"""
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return None
    # NFKC 归一化: 全角数字/字母/横线(１２３ｋｇ－)转半角, 中文Excel手输常见
    s = unicodedata.normalize("NFKC", str(text)).strip().lower().replace(" ", "")
    if not s:
        return None
    m = re.search(r"([\d.]+)\s*(mg|kg|lb|吨|磅|公斤|千克|克|[gt])?", s)
    if not m:
        return None
    try:
        val = float(m.group(1))
    except ValueError:
        return None
    unit = m.group(2) or "kg"
    return val * UNIT_FACTOR.get(unit, 1.0)

def parse_weight_range(text):
    """
    解析区间重量, 返回 (low, high) 单位 kg。
    支持: "1-2", "0.5~1.5", "1~2kg", "小于3", "3以上", "≥2", "<5", "2-"
    失败返回 None
    """
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return None
    # NFKC 归一化: 全角字符转半角
    s = unicodedata.normalize("NFKC", str(text)).strip().lower().replace(" ", "").replace("～", "~").replace("—", "-").replace("–", "-")
    if not s:
        return None

    # a-b / a~b / a至b
    m = re.search(r"^(\d+(?:\.\d+)?)[~\-至](\d+(?:\.\d+)?)", s)
    if m:
        lo = float(m.group(1)) * _unit_of(s)
        hi = float(m.group(2)) * _unit_of(s)
        return (min(lo, hi), max(lo, hi))

    # 小于/低于/< x
    m = re.search(r"(?:小于|低于|不超过|[<≤])\s*(\d+(?:\.\d+)?)", s)
    if m:
        v = float(m.group(1)) * _unit_of(s)
        return (0.0, v)

    # 大于/超过/>= x / 以上
    m = re.search(r"(?:大于|超过|高于|大于等于|[>≥])\s*(\d+(?:\.\d+)?)", s)
    if m:
        v = float(m.group(1)) * _unit_of(s)
        return (v, float("inf"))

    # 纯数字也当作区间 [x, x]
    m = re.fullmatch(r"(\d+(?:\.\d+)?)(mg|kg|lb|吨|磅|公斤|千克|克|[gt])?", s)
    if m:
        v = float(m.group(1)) * UNIT_FACTOR.get(m.group(2) or "kg", 1.0)
        return (v, v)
    return None

def _unit_of(s):
    if "公斤" in s or "千克" in s or "kg" in s:
        return 1.0
    if "克" in s:
        return 0.001
    for u in ("mg", "lb", "吨", "磅"):
        if u in s:
            return UNIT_FACTOR[u]
    if re.search(r"\d\s*t\b", s):
        return UNIT_FACTOR["t"]
    return 1.0  # 默认 kg

def weight_match(exact_kg, range_lr, tolerance=0.0):
    """判断具体重量是否落在区间内(可加容差), 返回 (bool, 说明)"""
    if exact_kg is None:
        return False, "A方重量无法解析"
    if range_lr is None:
        return False, "B方区间无法解析"
    lo, hi = range_lr
    if lo - tolerance <= exact_kg <= hi + tolerance:
        hi_s = "+inf" if hi == float("inf") else f"{hi:g}"
        lo_s = f"{lo:g}"
        return True, f"{exact_kg:.3f}kg ∈ [{lo_s}, {hi_s}]"
    hi_s = "+inf" if hi == float("inf") else f"{hi:g}"
    return False, f"{exact_kg:.3f}kg ∉ [{lo:g}, {hi_s}]"

# ---------- 主比对逻辑 ----------
def compare(file_a, file_b,
            key_a="快递单号", key_b="快递单号",
            weight_a="重量", weight_b="重量区间",
            sheet_a=0, sheet_b=0, tolerance=0.0):
    df_a = pd.read_excel(file_a, sheet_name=sheet_a, dtype={key_a: str})
    df_b = pd.read_excel(file_b, sheet_name=sheet_b, dtype={key_b: str})
    # 清洗单号: 去掉空值/空串, 避免空单元格被当作单号参与匹配或导致排序崩溃
    for df, key in ((df_a, key_a), (df_b, key_b)):
        df["_key"] = df[key].apply(lambda x: str(x).strip() if pd.notna(x) else "")
        df.dropna(subset=[key], inplace=True)
        df.drop(df[df["_key"] == ""].index, inplace=True)

    keys_a, keys_b = set(df_a["_key"]), set(df_b["_key"])
    only_a = keys_a - keys_b   # B 方缺失
    only_b = keys_b - keys_a   # A 方缺失
    common = keys_a & keys_b

    results = []  # dict: 类型/单号/A数据/B数据/说明

    # 业务规则: 同侧单号重复是重要对账线索(可能是\"一单多包裹\", 也可能是重复录入), 必须提示
    for side, df, other in (("A", df_a, "B"), ("B", df_b, "A")):
        vc = df["_key"].value_counts()
        for k, n in vc[vc > 1].items():
            results.append({
                "审查结果": f"提示-{side}方单号重复",
                "快递单号": k,
                "A方重量": f"{n}条记录" if side == "A" else None,
                "B方区间": f"{n}条记录" if side == "B" else None,
                "说明": f"{side}方该单号出现{n}次, 请人工确认是多包裹还是重复录入",
                "_row_a": None, "_row_b": None,
            })

    for k in sorted(common):
        rows_a = df_a[df_a["_key"] == k]
        rows_b = df_b[df_b["_key"] == k]
        n_a, n_b = len(rows_a), len(rows_b)
        for _, ra in rows_a.iterrows():
            for _, rb in rows_b.iterrows():
                exact = parse_weight_value(ra.get(weight_a))
                rng = parse_weight_range(rb.get(weight_b))
                ok, msg = weight_match(exact, rng, tolerance)
                results.append({
                    "审查结果": "正常" if ok else "重量不符",
                    "快递单号": k,
                    "A方重量": ra.get(weight_a),
                    "B方区间": rb.get(weight_b),
                    "A方记录数": n_a,
                    "B方记录数": n_b,
                    "说明": msg,
                    "_row_a": ra.name + 2,   # Excel 行号(含表头)
                    "_row_b": rb.name + 2,
                })

    for k in sorted(only_a):
        rows = df_a[df_a["_key"] == k]
        ra = rows.iloc[0]
        results.append({"审查结果": "异常-B方缺单", "快递单号": k,
                        "A方重量": ra.get(weight_a), "B方区间": None,
                        "A方记录数": len(rows), "B方记录数": 0,
                        "说明": "该单号仅在A方存在" + (f"(共{len(rows)}条)" if len(rows) > 1 else ""),
                        "_row_a": ra.name + 2, "_row_b": None})

    for k in sorted(only_b):
        rows = df_b[df_b["_key"] == k]
        rb = rows.iloc[0]
        results.append({"审查结果": "异常-A方缺单", "快递单号": k,
                        "A方重量": None, "B方区间": rb.get(weight_b),
                        "A方记录数": 0, "B方记录数": len(rows),
                        "说明": "该单号仅在B方存在" + (f"(共{len(rows)}条)" if len(rows) > 1 else ""),
                        "_row_a": None, "_row_b": rb.name + 2})

    return pd.DataFrame(results), df_a, df_b

# ---------- 输出: 标注原表 + 异常汇总 ----------
FILL_RED   = PatternFill("solid", fgColor="FFC7CE")
FONT_RED   = Font(color="9C0006", bold=True)

def _annotate_sheet(ws, df_rows, key_col, weight_col):
    """把 df_rows(原表数据+异常标记) 写入工作表并标红异常行"""
    cols = [c for c in df_rows.columns if not str(c).startswith("_")]
    note_col = len(cols) + 1
    for j, name in enumerate(cols, 1):
        c = ws.cell(row=1, column=j, value=name)
        c.font = Font(bold=True)
    hc = ws.cell(row=1, column=note_col, value="审查备注")
    hc.font = Font(bold=True)
    for i, (_, row) in enumerate(df_rows.iterrows(), 2):
        abnormal = bool(row.get("_异常"))
        for j, name in enumerate(cols, 1):
            v = row[name]
            if isinstance(v, float) and pd.isna(v):
                v = None
            c = ws.cell(row=i, column=j, value=v)
            if abnormal and name in (key_col, weight_col):
                c.fill, c.font = FILL_RED, FONT_RED
        if abnormal:
            c = ws.cell(row=i, column=note_col, value=row.get("_说明", ""))
            c.fill, c.font = FILL_RED, FONT_RED


def export_report(results, file_a, file_b, key_a, key_b,
                  weight_a, weight_b, out_path="审查报告.xlsx"):
    abnormal = results[results["审查结果"] != "正常"]
    summary = {
        "比对一致": int((results["审查结果"] == "正常").sum()),
        "重量不符": int((results["审查结果"] == "重量不符").sum()),
        "仅A方有(B缺单)": int((results["审查结果"] == "异常-B方缺单").sum()),
        "仅B方有(A缺单)": int((results["审查结果"] == "异常-A方缺单").sum()),
        "单号重复提示": int(results["审查结果"].str.startswith("提示-").sum()),
    }

    # 异常行在原表中的行号集合
    bad_a = {int(r) for r in abnormal["_row_a"].dropna()}
    bad_b = {int(r) for r in abnormal["_row_b"].dropna()}
    notes_a = {int(r["_row_a"]): f"[{r['审查结果']}] {r['说明']}"
               for _, r in abnormal.iterrows() if pd.notna(r["_row_a"])}
    notes_b = {int(r["_row_b"]): f"[{r['审查结果']}] {r['说明']}"
               for _, r in abnormal.iterrows() if pd.notna(r["_row_b"])}

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # 1) 全量明细 / 2) 异常提取
        results.drop(columns=["_row_a", "_row_b"]).to_excel(
            writer, sheet_name="比对明细", index=False)
        abnormal.drop(columns=["_row_a", "_row_b"]).to_excel(
            writer, sheet_name="异常提取", index=False)

        # 3) 标注后的 A 方副本
        wb = writer.book
        df_a = pd.read_excel(file_a, dtype={key_a: str})
        # 缺单单号的所有行都要标红(不能只标第一条)
        only_a_keys = set(results.loc[results["审查结果"] == "异常-B方缺单", "快递单号"])
        miss_a = df_a[key_a].apply(lambda x: str(x).strip() if pd.notna(x) else "").isin(only_a_keys)
        bad_a |= set(df_a.index[miss_a] + 2)
        notes_extra = {i + 2: "[异常-B方缺单] 该单号仅在A方存在"
                       for i in df_a.index[miss_a] if i + 2 not in notes_a}
        df_a["_异常"] = [(i + 2) in bad_a for i in range(len(df_a))]
        df_a["_说明"] = [notes_a.get(i + 2) or notes_extra.get(i + 2, "") for i in range(len(df_a))]
        ws = wb.create_sheet("A方标注")
        _annotate_sheet(ws, df_a, key_a, weight_a)

        # 4) 标注后的 B 方副本
        df_b = pd.read_excel(file_b, dtype={key_b: str})
        only_b_keys = set(results.loc[results["审查结果"] == "异常-A方缺单", "快递单号"])
        miss_b = df_b[key_b].apply(lambda x: str(x).strip() if pd.notna(x) else "").isin(only_b_keys)
        bad_b |= set(df_b.index[miss_b] + 2)
        notes_b.update({i + 2: "[异常-A方缺单] 该单号仅在B方存在"
                        for i in df_b.index[miss_b] if i + 2 not in notes_b})
        df_b["_异常"] = [(i + 2) in bad_b for i in range(len(df_b))]
        df_b["_说明"] = [notes_b.get(i + 2, "") for i in range(len(df_b))]
        ws = wb.create_sheet("B方标注")
        _annotate_sheet(ws, df_b, key_b, weight_b)

    return summary, out_path

if __name__ == "__main__":
    import sys
    print("用法示例:")
    print("  from compare_engine import compare, export_report")
    print("  results, *_ = compare('A.xlsx', 'B.xlsx')")
    print("  export_report(results, 'A.xlsx', 'B.xlsx', '快递单号','快递单号','重量','重量区间')")
